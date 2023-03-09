# TODO 导入openpyxl模块
import openpyxl

# TODO 使用openpyxl.Workbook()函数，创建一个新工作薄，赋值给newWb
newWb = openpyxl.Workbook()

# TODO 将默认的工作表Sheet赋值给aSheet变量
aSheet = newWb["Sheet"]

# TODO 将aSheet工作表名称修改为“七年级1班课表”
aSheet.title = "七年级1班课表"

# TODO 使用for循环和range，逐个遍历2~10的数字
for i in range(2,11):
    # TODO 利用格式化字符串"七年级x班课表",创建不同班级的工作表
    newWb.create_sheet(f"七年级{i}班课表")

# TODO 使用for循环遍历工作簿对象的worksheets属性
for sheet in newWb.worksheets:

    # TODO 给每一个工作表设置表头，B1单元格设置为星期一，C1单元格设置为星期二
    sheet["B1"].value = "星期一"
    sheet["C1"].value = "星期二"
    sheet["D1"].value = "星期三"
    sheet["E1"].value = "星期四"
    sheet["F1"].value = "星期五"

    # TODO 使用merge_cells()函数，合并单元格A2:A4
    sheet.merge_cells("A2:A4")
    # TODO 将单元格A2的值设置为上午
    sheet["A2"].value = "上午"

    # TODO 使用merge_cells()函数，合并单元格A5:A6
    sheet.merge_cells("A5:A6")
    # TODO 将单元格A5的值设置为下午
    sheet["A5"].value = "下午"

# TODO 将工作簿保存到路径 /Users/yequ/Desktop/七年级课表.xlsx
newWb.save("d:\\test\\七年级课表2.xlsx")