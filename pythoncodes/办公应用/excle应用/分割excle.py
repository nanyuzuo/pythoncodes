# 阿珍需要将网店2019年每个月份的销售数据总表拆分成按商品名称分类的分表，便于查看每种商品的销售数据。12个月份的销售订单路径：/Users/zhener/doc
# 每个订单命名方式为：2019年X月销售订单.xlsx
# 每个商品的销售数据存储在销售订单数据工作表中
# TODO 使用import导入openpyxl
import openpyxl

# TODO 使用import导入os模块
import os

# TODO 将销售订单路径'/Users/zhener/doc'赋值给path
path = "/Users/zhener/doc"

# TODO 使用os.listdir()获取文件中的所有文件名，赋值给fileNames
fileNames = os.listdir(path)

# TODO for循环遍历列表中的所有文件名
for file in fileNames:

    # TODO 使用os.path.join()函数拼接文件路径，赋值给filePath
    filePath = os.path.join(path,file)

    # TODO 使用openpyxl.load_workbook()打开文档赋值给wb
    wb = openpyxl.load_workbook(filePath)

    # TODO 将文档中的 "销售订单数据" 工作表赋值给orderSheet
    orderSheet = wb["销售订单数据"]

    # TODO 定义一个用于存储商品名称的空列表wbName,不可以用list直接作为变量名
    wbName = []

    # TODO for循环遍历销售订单数据工作表的 .rows 属性
    for row in orderSheet.rows:

        # TODO 取第3列的商品名称赋值给name
        name = row[2].value

        # TODO 使用if判断商品名称不在wbName列表中
        if name not in wbName:

            # TODO 使用append()函数将商品名追加到列表中
            wbName.append(name)

    # TODO 使用for循环从第2个数据开始遍历wbName列表，第1个是"商品名"，也就是列表切片[1:]
    for productName in wbName[1:]:

        # TODO 使用create_sheet()创建新工作表，并将商品名作为工作表名称，赋值给newSheet
        newSheet = wb.create_sheet(productName)

        # TODO 按照"销售订单数据"工作表写入表头信息，在A1的表格中填写"订单号"
        newSheet["A1"].value = "订单号"
        newSheet["B1"].value = "商品ID"
        newSheet["C1"].value = "商品名称"
        newSheet["D1"].value = "品牌"
        newSheet["E1"].value = "类别"
        newSheet["F1"].value = "规格"
        newSheet["G1"].value = "单价"
        newSheet["H1"].value = "数量"
        newSheet["I1"].value = "总价"

    # TODO 保存工作簿到原路径filePath
    wb.save(filePath)