# 使用from...import从pyecharts.charts中导入Graph
from pyecharts.charts import Graph
# 使用from...import从pyecharts中导入options简写为opts
from pyecharts import options as opts
# 使用import导入openpyxl模块
import openpyxl

# 文件路径/Users/bing/基本信息.xlsx 赋值给path
path = "/Users/bing/基本信息.xlsx"
# 使用openpyxl.load_workbook()读文档赋值给wb
wb = openpyxl.load_workbook(path)
# 使用wb[]读取工作表部门，赋值给depart
depart = wb["部门"]
# 使用.max_row获取最大行数，赋值给depart_row
depart_row = depart.max_row

# 新建列表category
category = []
# 使用for循环配合range()从第2行开始遍历所有行数据
for row in range(2, depart_row+1):
    # 新建字典cate_dict
    cate_dict = {}
    # 使用depart[]读取行数据赋值给cate_data
    cate_data = depart[row]
    # 使用.value将cate_data第1个元素，赋值给cate_dict["name"]
    cate_dict["name"] = cate_data[0].value
    # 使用append()将字典追加到列表category
    category.append(cate_dict)

# 使用wb[]读取工作表合作关系，赋值给partner
partner = wb["合作关系"]
# 新建字典name_dict
name_dict = {}
# 新建列表links_list
links_list = []
# 使用for循环遍历partner工作表的所有行数据
for par_row in range(1,partner.max_row+1):
    # 使用partner[]读取行数，赋值给data
    data = partner[par_row]
    # 使用.value将data的第1个元素，赋值给name_first
    name_first = data[0].value
    # 判断名字不在字典的键中时
    if name_first not in name_dict.keys():
        # 将name_first作为键，1作为值，写入字典name_dict中
        name_dict[name_first] = 1
    # 其他情况
    else:
        # 字典name_dict的值计数加1
        name_dict[name_first] += 1

    # 使用.value将data的第2个元素，赋值给name_second
    name_second = data[1].value
    # 使用split()传入逗号对name_second进行分割，赋值给name_second_list
    name_second_list = name_second.split(",")
    # 使用for循环遍历name_second_list的每项
    for name_second_item in name_second_list:
        # 判断名字不在字典的键中时
        if name_second_item not in name_dict.keys():
            # 将name_second_item作为键，1作为值，写入字典name_dict中
            name_dict[name_second_item] = 1
        # 其他情况
        else:
            # 字典name_dict的值计数加1
            name_dict[name_second_item] += 1

        # 新建字典temp_dict
        temp_dict = {}
        # 将source作为键，name_first作为值写入字典temp_dict
        temp_dict["source"] = name_first
        # 将target作为键，name_second_item作为值写入字典temp_dict
        temp_dict["target"] = name_second_item
        # 将字典的内容追加到列表links_list中
        links_list.append(temp_dict)

# 使用wb[]读取个人工作表，赋值给person
person = wb["个人"]
# 新建列表nodes_list
nodes_list = []
# 使用for循环配合range()从第2行开始遍历所有行数据
for per_row in range(2, person.max_row+1):
    # 使用person[]读取行数，赋值给per_data
    per_data = person[per_row]
    # 新建字典nodes_dict
    nodes_dict = {}
    # 读取第一个元素，写入字典nodes_dict
    nodes_dict["id"] = per_data[0].value
    # 读取第一个元素，写入字典nodes_dict
    nodes_dict["name"] = per_data[0].value
    # 读取第二个元素，写入字典nodes_dict
    nodes_dict["category"] = per_data[1].value
    # 将per_data[0].value传入字典name_dict取对应的值，写入字典nodes_dict
    nodes_dict["symbolSize"] = name_dict[per_data[0].value]
    # 使用append()将nodes_dict追加到列表nodes_list中
    nodes_list.append(nodes_dict)

# 使用Graph()函数创建对象赋值给graph
graph = Graph()

# 调用add()函数，设置series_name为空
# 将category赋值给categories
# 将nodes_list赋值给nodes，将links_list赋值给links
# 设置layout="circular"
# 设置is_rotate_label=True
# 调用LineStyleOpts()函数，传入参数color="source"，curve=0.3
graph.add(
    series_name="",
    categories=category,
    nodes=nodes_list,
    links=links_list,
    layout="circular",
    is_rotate_label=True,
    linestyle_opts=opts.LineStyleOpts(color="source", curve=0.3)
    )
# 将参数orient="vertical", pos_left="3%", pos_top="20%"传入LegendOpts()
# 将legend_opts=opts.LegendOpts()传入set_global_opts()
# TODO 使用TitleOpts()，设置标题为"员工合作关系图"，赋值给title_opts
graph.set_global_opts(
    legend_opts=opts.LegendOpts(orient="vertical", pos_left="3%", pos_top="20%"),
    title_opts=opts.TitleOpts(title="员工合作关系图")
)

# 使用render()生成文件存储/Users/bing/graph.html
graph.render("/Users/bing/graph.html")