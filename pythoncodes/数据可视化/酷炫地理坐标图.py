# 使用from...import...从pyecharts.charts导入Geo
from pyecharts.charts import Geo
# 使用from...import...从pyecharts中导入options，并简写为opts
from pyecharts import options as opts
# 使用import导入openpyxl模块
import openpyxl
# 使用from...import...从pyecharts.commons.utils中导入JsCode
from pyecharts.commons.utils import JsCode

# 使用openpyxl.load_workbook()读取文件，并赋值给wb
wb = openpyxl.load_workbook("/Users/JJ/travel_routes.xlsx")

# 使用中括号读取工作表"travel_route"，并赋值给sheet
sheet = wb["travel_route"]

# 新建列表routeList
routeList = []

# 使用for循环和range()函数，依次遍历工作表中的所有行数据
for row in range(2, sheet.max_row+1):
    
    # 将当前行数据赋值给变量routeInfo
    routeInfo = sheet[row]

    # 获取routeInfo的第一项的值和第二项的值
    # 以元组的格式组合，并赋值给data
    data = (routeInfo[0].value, routeInfo[1].value)

    # 使用append()函数将data添加进routeList
    routeList.append(data)

# 定义渐变色，并赋值给变量line_color_js
line_color_js = """
new echarts.graphic.LinearGradient(
    0,0,0,1,
    [{offset: 0,color: '#42aecc'},{offset: 1,color: '#f5602c'}])
"""

# 使用Geo()函数创建Geo对象，把背景颜色设置为"#475262"，并将结果赋值给变量geo
geo = Geo(
    init_opts=opts.InitOpts(bg_color="#475262")
)

# 使用add_schema()函数，将地图类型设置为"china"，地图底色改为"#2d3948"，描边颜色改为"#58667a"
# 高亮状态下颜色改为"#2a333d"
geo.add_schema(
    maptype="china",
    itemstyle_opts=opts.ItemStyleOpts(
            color="#2d3948", 
            border_color="#58667a"
    ),
    emphasis_itemstyle_opts=opts.ItemStyleOpts(
            color="#2a333d"
    )
)

# TODO 使用add()函数、linestyle_opts参数和线样式基本配置项
# 将线条颜色改为line_color_js
geo.add(
    series_name="", 
    data_pair=routeList, 
    type_="lines",
    label_opts=opts.LabelOpts(is_show=False),
    symbol="circle",
    symbol_size=8,
    linestyle_opts=opts.LineStyleOpts(
        opacity=0.1,curve=0.1,
        color=JsCode(line_color_js)
    )
)

# 使用render()函数生成地理坐标系图，并存储到路径 "/Users/JJ/geo.html" 下
geo.render("/Users/JJ/geo.html")