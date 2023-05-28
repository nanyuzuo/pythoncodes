# 使用from...import...从pyecharts.charts导入Geo
from pyecharts.charts import Geo
# TODO 使用from...import...从pyecharts中导入options，并简写为opts
from pyecharts import options as opts

# 使用import导入openpyxl模块
import openpyxl

# 使用openpyxl.load_workbook()读取文件，并赋值给wb
wb = openpyxl.load_workbook("/Users/JJ/travel_routes.xlsx")

# 使用中括号读取工作表"travel_route"，并赋值给sheet
sheet = wb["travel_route"]

# 新建列表routeList
routeList = []

# 使用for循环和range()函数，依次遍历工作表中的所有行数据
for row in range(2, sheet.max_row+1):
    
    # 将当前行数据赋值给变量routeInfo
    routeInfo = sheet[row]

    # 获取routeInfo的第一项的值和第二项的值
    # 以元组的格式组合，并赋值给data
    data = (routeInfo[0].value, routeInfo[1].value)

    # 使用append()函数将data添加进routeList
    routeList.append(data)

# 使用Geo()函数创建Geo对象，并赋值给变量geo
geo = Geo()

# 使用add_schema()函数和maptype参数，将地图类型设置为"china"
geo.add_schema(
    maptype="china"
)

# TODO 使用label_opts参数隐藏标签
geo.add(
    series_name="", 
    data_pair=routeList, 
    type_="lines",
    label_opts=opts.LabelOpts(is_show=False)
)

# 使用render()函数生成地理坐标系图，并存储到路径 "/Users/JJ/geo.html" 下
geo.render("/Users/JJ/geo.html")